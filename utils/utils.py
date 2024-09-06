
import time

from openpyxl import Workbook, load_workbook


# достаем refer_id из команды /start
def get_refer_id(command_args):
    try:
        return int(command_args)
    except (TypeError, ValueError):
        return None


# формируем файл с отчетом по пользователям
def get_statistic_file(all_users_data: list[dict]) -> str:

    timestr = time.strftime('%Y.%m.%d_%H-%M')
    name_file = f'referral_USERS_VB({timestr}).xlsx'
    count_row = 2
    result_number = 1

    wb = Workbook()
    wb.save(name_file)
    wb = load_workbook(name_file)
    ws = wb.active

    ws.cell(row=1, column=1, value='№')
    ws.cell(row=1, column=2, value='user_id')
    ws.cell(row=1, column=3, value='full_name')
    ws.cell(row=1, column=4, value='user_login')
    ws.cell(row=1, column=5, value='кто реферал')
    ws.cell(row=1, column=6, value='сколько привлек')
    ws.cell(row=1, column=7, value='на сколько каналов подписан')
    ws.cell(row=1, column=8, value='добавлен в базу')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 20

    wb.save(name_file)

    for user in all_users_data:
        ws.cell(row=count_row, column=1, value=f'{result_number}')
        ws.cell(row=count_row, column=2, value=user['user_id'])
        ws.cell(row=count_row, column=3, value=user['full_name'] if user['full_name'] else None)
        ws.cell(row=count_row, column=4, value=f'@{user["user_login"]}' if user['user_login'] else None)
        ws.cell(row=count_row, column=5, value=user['refer_id'] if user['refer_id'] else 'None')
        ws.cell(row=count_row, column=6, value=user['count_refer'])
        ws.cell(row=count_row, column=7, value=user['chat_member_counter'])
        ws.cell(row=count_row, column=8, value=user['date_reg'])
        count_row += 1
        result_number += 1

    wb.save(name_file)

    return name_file
